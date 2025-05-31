
from openai import OpenAI
import pandas as pd
from openpyxl import Workbook
from openpyxl import load_workbook
import os
# 读取CSV文件
df = pd.read_excel('***.xlsx')

# 设置OpenAI客户端
client = OpenAI(
    base_url='***',
    api_key='sk-***',
)


# 定义一个函数来分析代码并返回结果
def analyze_code(code):
    messages = [
        {
            "role": "user",
            "content": "Task Desctiption & Instructions "
                       "You are a software security engineer. Analyze the information about the code snippet provided below. "
                       "First, analyze the exploitability of the code, including the necessary conditions and ways to use it."
                       "Secondly, the impact analysis of the vulnerability is carried out to give the direct consequences of the successful exploitation of the vulnerability."
                       "Finally, give the scope of impact after the vulnerability is exploited. "
                       "Output Format Specification"
                       "Desired format: Exploitability: '<necessary condition>' and '<way>'"
                       "Impact: '<impact>'"
                       "Scope: '<scope>'"
                       "Within 100 words."
                       "Code Placeholder Code Snippet: ```c {} ```".format(code),
        }
    ]

    response = client.chat.completions.create(
        messages=messages,
        model="gpt-3.5-turbo",
        # model="deepseek-chat",
    )
    for message in response.choices:
        print(message.message.content)
    result = {}
    for message in response.choices:
        for line in message.message.content.split('\n'):
            if line.startswith('Exploitability:'):
                result['exploitability'] = line[len('Exploitability: '):].strip()
            elif line.startswith('Impact:'):
                result['impact'] = line[len('Impact: '):].strip()
            elif line.startswith('Scope:'):
                result['scope'] = line[len('Scope: '):].strip()
    return result
# 创建一个新的DataFrame来存储结果
results_df = pd.DataFrame(columns=['func_before', 'exploitability', 'impact', 'scope'])
count = 0
start_index = 0
excel_filename = '***.xlsx'

# 遍历DataFrame，分析代码并将结果逐条写入Excel文件
for index, row in df.iloc[start_index:].iterrows():
    result = analyze_code(row['func_before'])
    new_row = [
        row['func_before'],
        result.get('exploitability', ''),
        result.get('impact', ''),
        result.get('scope', ''),
    ]
    # 如果文件不存在，创建一个新的工作簿并写入表头和数据
    if not os.path.exists(excel_filename):
        wb = Workbook()
        ws = wb.active
        ws.append(['func_before', 'exploitability', 'impact', 'scope'])
        ws.append(new_row)
        wb.save(excel_filename)
    else:
        wb = load_workbook(excel_filename)
        ws = wb.active
        ws.append(new_row)
        wb.save(excel_filename)
    count += 1
    print("处理完成{}个啦".format(count))